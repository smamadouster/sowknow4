#!/usr/bin/env python3
"""
Standalone spreadsheet extractor for subprocess-based timeout isolation.

Called by TextExtractor via ``subprocess.run(timeout=120)`` to prevent
openpyxl/xlrd hangs from blocking Celery workers.

Usage:
    python -m app.services._spreadsheet_extractor <xlsx|xls> <file_path>
"""

import json
import sys


def extract_xlsx(file_path: str) -> dict:
    """Extract text from XLSX using openpyxl with safety bounds."""
    from openpyxl import load_workbook

    _MAX_SHEETS = 20
    _MAX_ROWS_PER_SHEET = 5000
    _MAX_TOTAL_CHARS = 500_000
    _ROW_BATCH_SIZE = 50

    wb = load_workbook(file_path, read_only=True)
    try:
        text_parts = []
        sheets_processed = 0
        total_chars = 0
        truncated = False

        for sheet_name in wb.sheetnames:
            if sheets_processed >= _MAX_SHEETS:
                break

            sheet = wb[sheet_name]
            # Skip hidden/very-hidden sheets (defensive: getattr for read-only quirks)
            if getattr(sheet, "sheet_state", "visible") != "visible":
                continue

            sheets_processed += 1
            sheet_text_parts = [f"[Sheet: {sheet_name}]"]
            batch_rows: list[str] = []
            row_count = 0

            # max_row limits the iterator so openpyxl doesn't walk huge declared dimensions
            for row in sheet.iter_rows(values_only=True, max_row=_MAX_ROWS_PER_SHEET):
                row_count += 1
                if row_count > _MAX_ROWS_PER_SHEET:
                    sheet_text_parts.append("[... rows truncated at 5000 ...]")
                    break

                # Skip rows where all cells are None or whitespace-only
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue

                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                batch_rows.append(row_text)

                if len(batch_rows) >= _ROW_BATCH_SIZE:
                    batch_text = "\n".join(batch_rows)
                    if total_chars + len(batch_text) > _MAX_TOTAL_CHARS:
                        truncated = True
                        break
                    sheet_text_parts.append(batch_text)
                    total_chars += len(batch_text) + 2  # +2 for \n\n
                    batch_rows = []

            if not truncated and batch_rows:
                batch_text = "\n".join(batch_rows)
                if total_chars + len(batch_text) <= _MAX_TOTAL_CHARS:
                    sheet_text_parts.append(batch_text)
                    total_chars += len(batch_text) + 2
                else:
                    truncated = True

            if len(sheet_text_parts) > 1:
                text_parts.append("\n\n".join(sheet_text_parts))

            if truncated:
                text_parts.append("[... document truncated at 500,000 characters ...]")
                break

        return {
            "text": "\n\n".join(text_parts),
            "pages": sheets_processed,
            "source": "openpyxl",
        }
    finally:
        wb.close()


def extract_xls(file_path: str) -> dict:
    """Extract text from legacy XLS using xlrd with safety bounds."""
    import xlrd

    _MAX_SHEETS = 20
    _MAX_ROWS_PER_SHEET = 5000
    _MAX_TOTAL_CHARS = 500_000
    _ROW_BATCH_SIZE = 50

    wb = xlrd.open_workbook(file_path, on_demand=True)
    try:
        text_parts = []
        sheets_processed = 0
        total_chars = 0
        truncated = False

        for sheet_idx in range(min(wb.nsheets, _MAX_SHEETS)):
            sheet = wb.sheet_by_index(sheet_idx)
            sheets_processed += 1
            sheet_text_parts = [f"[Sheet: {sheet.name}]"]
            batch_rows: list[str] = []
            row_count = min(sheet.nrows, _MAX_ROWS_PER_SHEET)

            for row_idx in range(row_count):
                row_values = sheet.row_values(row_idx)
                # Skip rows where all cells are empty/whitespace
                if not any(cell is not None and str(cell).strip() for cell in row_values):
                    continue

                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row_values])
                batch_rows.append(row_text)

                if len(batch_rows) >= _ROW_BATCH_SIZE:
                    batch_text = "\n".join(batch_rows)
                    if total_chars + len(batch_text) > _MAX_TOTAL_CHARS:
                        truncated = True
                        break
                    sheet_text_parts.append(batch_text)
                    total_chars += len(batch_text) + 2
                    batch_rows = []

            if not truncated and batch_rows:
                batch_text = "\n".join(batch_rows)
                if total_chars + len(batch_text) <= _MAX_TOTAL_CHARS:
                    sheet_text_parts.append(batch_text)
                    total_chars += len(batch_text) + 2
                else:
                    truncated = True

            if len(sheet_text_parts) > 1:
                text_parts.append("\n\n".join(sheet_text_parts))

            if truncated:
                text_parts.append("[... document truncated at 500,000 characters ...]")
                break

        return {
            "text": "\n\n".join(text_parts),
            "pages": sheets_processed,
            "source": "xlrd",
        }
    finally:
        wb.release_resources()


if __name__ == "__main__":
    fmt = sys.argv[1]
    path = sys.argv[2]
    try:
        if fmt == "xlsx":
            result = extract_xlsx(path)
        elif fmt == "xls":
            result = extract_xls(path)
        else:
            result = {"text": "", "error": f"Unknown format: {fmt}", "pages": 0}
    except AttributeError as exc:
        # xlrd 1.2.0 triggers getiterator() which was removed in Python 3.9+.
        # Surface a cleaner error so the caller can fall back to openpyxl.
        err_msg = str(exc)
        if "getiterator" in err_msg:
            err_msg = "xlrd parser incompatible with this file on Python 3.11+"
        result = {"text": "", "error": err_msg, "pages": 0}
    except Exception as exc:
        result = {"text": "", "error": str(exc), "pages": 0}
    print(json.dumps(result))
