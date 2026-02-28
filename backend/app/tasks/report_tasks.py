"""
Celery tasks for asynchronous report generation and cleanup.

Tasks:
  1. generate_pdf_report   — generate a PDF analytics report
  2. generate_excel_export — generate an Excel data export
  3. cleanup_old_reports   — delete report files older than N days
"""

from __future__ import annotations

import logging
import os
import time
from datetime import datetime

from celery import shared_task

from app.tasks.base import base_task_failure_handler, store_dlq_on_max_retries

logger = logging.getLogger(__name__)

# Directory where generated reports are stored (bind-mounted in Docker)
REPORTS_DIR = os.getenv("REPORTS_DIR", "/tmp/sowknow_reports")  # nosec B108


def _ensure_reports_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Task 1: generate_pdf_report
# ---------------------------------------------------------------------------


@shared_task(
    bind=True,
    name="app.tasks.report_tasks.generate_pdf_report",
    queue="celery",
    autoretry_for=(Exception,),
    retry_kwargs={"max_retries": 2},
    retry_backoff=True,
)
def generate_pdf_report(
    self,
    report_type: str,
    filters: dict,
    user_id: str,
    output_filename: str | None = None,
) -> dict:
    """
    Generate a PDF report asynchronously.

    Args:
        report_type:     One of "anomaly", "usage", "document_stats".
        filters:         Query filters (e.g. {"date_from": "...", "date_to": "..."}).
        user_id:         UUID of the requesting user (for audit trail).
        output_filename: Optional filename; auto-generated if not provided.

    Returns:
        {
            "status": "completed",
            "report_type": str,
            "file_path": str,
            "file_size_bytes": int,
            "generation_time_seconds": float
        }
    """
    _ensure_reports_dir()
    start = time.time()

    if output_filename is None:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{report_type}_{ts}_{self.request.id[:8]}.pdf"

    output_path = os.path.join(REPORTS_DIR, output_filename)

    try:
        _generate_pdf_content(report_type, filters, output_path)

        file_size = os.path.getsize(output_path)
        duration = round(time.time() - start, 2)

        logger.info(f"generate_pdf_report: {report_type} completed in {duration}s ({file_size} bytes) → {output_path}")
        return {
            "status": "completed",
            "report_type": report_type,
            "file_path": output_path,
            "file_size_bytes": file_size,
            "generation_time_seconds": duration,
        }

    except Exception as exc:
        logger.error(f"generate_pdf_report ({report_type}) failed: {exc}")
        store_dlq_on_max_retries(
            self,
            exc,
            extra_metadata={"report_type": report_type, "user_id": user_id},
        )
        raise


def _generate_pdf_content(report_type: str, filters: dict, output_path: str):
    """
    Build PDF content for the requested report type.

    Uses fpdf2 when available, falls back to a plain-text stub so the task
    does not crash in environments without fpdf2.
    """
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        try:
            from fpdf import FPDF  # type: ignore

            _write_pdf_with_fpdf(FPDF, report_type, filters, output_path, db)
        except ImportError:
            # Fallback: write a text file with .pdf extension
            _write_text_stub(report_type, filters, output_path, db)
    finally:
        db.close()


def _write_pdf_with_fpdf(FPDF, report_type: str, filters: dict, output_path: str, db):
    """Generate a real PDF using fpdf2."""
    from sqlalchemy import func

    from app.models.document import Document

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=14)
    pdf.cell(0, 10, f"SOWKNOW Report: {report_type.upper()}", ln=True)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().isoformat()} UTC", ln=True)
    pdf.cell(0, 8, f"Filters: {filters}", ln=True)
    pdf.ln(5)

    if report_type == "document_stats":
        counts = db.query(Document.status, func.count(Document.id)).group_by(Document.status).all()
        for status, count in counts:
            pdf.cell(0, 7, f"  {status}: {count}", ln=True)

    pdf.output(output_path)


def _write_text_stub(report_type: str, filters: dict, output_path: str, db):
    """Fallback: write minimal content when fpdf2 is not installed."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"SOWKNOW Report — {report_type}\n")
        f.write(f"Generated: {datetime.utcnow().isoformat()} UTC\n")
        f.write(f"Filters: {filters}\n")
        f.write("(fpdf2 not installed — plain-text fallback)\n")


# ---------------------------------------------------------------------------
# Task 2: generate_excel_export
# ---------------------------------------------------------------------------


@shared_task(
    bind=True,
    name="app.tasks.report_tasks.generate_excel_export",
    queue="celery",
    autoretry_for=(Exception,),
    retry_kwargs={"max_retries": 2},
    retry_backoff=True,
)
def generate_excel_export(
    self,
    export_type: str,
    filters: dict,
    user_id: str,
    output_filename: str | None = None,
) -> dict:
    """
    Generate an Excel export asynchronously.

    Args:
        export_type:     One of "documents", "users", "audit_log".
        filters:         Query filters.
        user_id:         UUID of the requesting user.
        output_filename: Optional filename; auto-generated if not provided.

    Returns:
        {
            "status": "completed",
            "export_type": str,
            "file_path": str,
            "file_size_bytes": int,
            "row_count": int,
            "generation_time_seconds": float
        }
    """
    _ensure_reports_dir()
    start = time.time()

    if output_filename is None:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{export_type}_{ts}_{self.request.id[:8]}.xlsx"

    output_path = os.path.join(REPORTS_DIR, output_filename)

    try:
        row_count = _generate_excel_content(export_type, filters, output_path)
        file_size = os.path.getsize(output_path)
        duration = round(time.time() - start, 2)

        logger.info(
            f"generate_excel_export: {export_type} completed in {duration}s ({row_count} rows, {file_size} bytes)"
        )
        return {
            "status": "completed",
            "export_type": export_type,
            "file_path": output_path,
            "file_size_bytes": file_size,
            "row_count": row_count,
            "generation_time_seconds": duration,
        }

    except Exception as exc:
        logger.error(f"generate_excel_export ({export_type}) failed: {exc}")
        store_dlq_on_max_retries(
            self,
            exc,
            extra_metadata={"export_type": export_type, "user_id": user_id},
        )
        raise


def _generate_excel_content(export_type: str, filters: dict, output_path: str) -> int:
    """Build Excel content, returns row count."""
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        try:
            import openpyxl  # type: ignore

            return _write_excel_with_openpyxl(openpyxl, export_type, filters, output_path, db)
        except ImportError:
            return _write_excel_stub(export_type, output_path, db)
    finally:
        db.close()


def _write_excel_with_openpyxl(openpyxl, export_type, filters, output_path, db) -> int:
    """Write Excel using openpyxl."""
    from app.models.audit import AuditLog
    from app.models.document import Document

    wb = openpyxl.Workbook()
    ws = wb.active
    row_count = 0

    if export_type == "documents":
        ws.title = "Documents"
        ws.append(["ID", "Filename", "Status", "Bucket", "Size", "Created At"])
        for doc in db.query(Document).all():
            ws.append(
                [
                    str(doc.id),
                    doc.filename,
                    doc.status.value if doc.status else "",
                    doc.bucket.value if doc.bucket else "",
                    doc.size,
                    doc.created_at.isoformat() if doc.created_at else "",
                ]
            )
            row_count += 1

    elif export_type == "audit_log":
        ws.title = "Audit Log"
        ws.append(["ID", "User ID", "Action", "Resource Type", "Created At"])
        for entry in db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(10000).all():
            ws.append(
                [
                    str(entry.id),
                    str(entry.user_id),
                    entry.action.value if entry.action else "",
                    entry.resource_type or "",
                    entry.created_at.isoformat() if entry.created_at else "",
                ]
            )
            row_count += 1

    wb.save(output_path)
    return row_count


def _write_excel_stub(export_type: str, output_path: str, db) -> int:
    """Fallback when openpyxl is not installed."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"SOWKNOW Export — {export_type}\n")
        f.write("(openpyxl not installed — plain-text fallback)\n")
    return 0


# ---------------------------------------------------------------------------
# Task 3: cleanup_old_reports
# ---------------------------------------------------------------------------


@shared_task(
    bind=True,
    name="app.tasks.report_tasks.cleanup_old_reports",
    queue="celery",
)
def cleanup_old_reports(self, days_to_keep: int = 7) -> dict:
    """
    Delete generated report files older than *days_to_keep* days.

    Scheduled by Celery Beat at 02:00 AM daily.

    Args:
        days_to_keep: Files older than this many days are deleted.

    Returns:
        {"status": "completed", "deleted": int, "errors": int}
    """
    _ensure_reports_dir()
    cutoff = time.time() - days_to_keep * 86400
    deleted = 0
    errors = 0

    try:
        for fname in os.listdir(REPORTS_DIR):
            fpath = os.path.join(REPORTS_DIR, fname)
            try:
                if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                    os.remove(fpath)
                    deleted += 1
            except Exception as file_err:
                logger.warning(f"cleanup_old_reports: could not delete {fpath}: {file_err}")
                errors += 1

        logger.info(f"cleanup_old_reports: deleted {deleted} files older than {days_to_keep} days")
        return {"status": "completed", "deleted": deleted, "errors": errors}

    except Exception as exc:
        logger.error(f"cleanup_old_reports failed: {exc}")
        raise


# ─────────────────────────────────────────────────────────────────────────────
# on_failure callback
# ─────────────────────────────────────────────────────────────────────────────


def on_generate_pdf_report_failure(self, exc, task_id, args, kwargs, einfo) -> None:
    """on_failure callback for the generate_pdf_report task."""
    logger.error(f"on_generate_pdf_report_failure: task_id={task_id} exc={exc!r}")
    base_task_failure_handler(
        task_self=self,
        exception=exc,
        task_id=task_id,
        args=args,
        kwargs=kwargs,
        traceback=einfo,
        is_critical=False,
    )
